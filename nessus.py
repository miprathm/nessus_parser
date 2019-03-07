import os
import re
import openpyxl
import sys
import xml.etree.ElementTree as ET
import openpyxl

pathname = os.path.abspath(sys.argv[0])
if len(sys.argv) > 1 :
	pathname = os.path.abspath(sys.argv[1])
print(pathname)


#!/usr/bin/python
f = open(pathname, 'r')
xml_content = f.read()
f.close()

wb = openpyxl.load_workbook("ips.xlsx")
sheet = wb.get_active_sheet()

current_ip = 2;
while sheet.cell(row=current_ip,column=1).value is not None:
	current_ip += 1
print("Processing "+pathname)
print(current_ip)
vulnerabilities = {}
single_params = ["agent", "cvss3_base_score", "cvss3_temporal_score", "cvss3_temporal_vector", "cvss3_vector",
                 "cvss_base_score", "cvss_temporal_score", "cvss_temporal_vector", "cvss_vector", "description",
                 "exploit_available", "exploitability_ease", "exploited_by_nessus", "fname", "in_the_news",
                 "patch_publication_date", "plugin_modification_date", "plugin_name", "plugin_publication_date",
                 "plugin_type", "script_version", "see_also", "solution", "synopsis", "vuln_publication_date"]
root = ET.fromstring(xml_content)
for block in root:
    if block.tag == "Report":
        for report_host in block:
            host_properties_dict = dict()
            for report_item in report_host:
                if report_item.tag == "HostProperties":
                    for host_properties in report_item:
                        host_properties_dict[host_properties.attrib['name']] = host_properties.text
            for report_item in report_host:
                if 'pluginName' in report_item.attrib:
                    vulner_id = report_host.attrib['name'] + "|" + report_item.attrib['port'] + "|"  + \
                                report_item.attrib['protocol'] + "|" + report_item.attrib['pluginID']
                    if not vulner_id in vulnerabilities:
                        vulnerabilities[vulner_id] = dict()
                    vulnerabilities[vulner_id]['port'] = report_item.attrib['port']
                    vulnerabilities[vulner_id]['pluginName'] = report_item.attrib['pluginName']
                    vulnerabilities[vulner_id]['pluginFamily'] = report_item.attrib['pluginFamily']
                    vulnerabilities[vulner_id]['pluginID'] = report_item.attrib['pluginID']
                    vulnerabilities[vulner_id]['svc_name'] = report_item.attrib['svc_name']
                    vulnerabilities[vulner_id]['protocol'] = report_item.attrib['protocol']
                    vulnerabilities[vulner_id]['severity'] = report_item.attrib['severity']
                    for param in report_item:
                        if param.tag == "risk_factor":
                            risk_factor = param.text
                            vulnerabilities[vulner_id]['host'] = report_host.attrib['name']
                            vulnerabilities[vulner_id]['riskFactor'] = risk_factor
                        elif param.tag == "plugin_output":
                            if not "plugin_output" in vulnerabilities[vulner_id]:
                                vulnerabilities[vulner_id]["plugin_output"] = list()
                            if not param.text in vulnerabilities[vulner_id]["plugin_output"]:
                                vulnerabilities[vulner_id]["plugin_output"].append(param.text)
                        else:
                            if not param.tag in single_params:
                                if not param.tag in vulnerabilities[vulner_id]:
                                    vulnerabilities[vulner_id][param.tag] = list()
                                if not isinstance(vulnerabilities[vulner_id][param.tag], list):
                                    vulnerabilities[vulner_id][param.tag] = [vulnerabilities[vulner_id][param.tag]]
                                if not param.text in vulnerabilities[vulner_id][param.tag]:
                                    vulnerabilities[vulner_id][param.tag].append(param.text)
                            else:
                                vulnerabilities[vulner_id][param.tag] = param.text
                    for param in host_properties_dict:
                        vulnerabilities[vulner_id][param] = host_properties_dict[param]
keys = vulnerabilities.keys()
ssl_cipher_plugin_ids = ['26928','65821']
ssl_cipher_finder = re.compile(r'(\w{3,4}\-\w{3,4}(\-\w{3,4})*?(\(.+\))?)\s*Kx=.*\s*Au=.*\s*Enc=.*\s*Mac=')

for key in keys:
	name, port, protocol, pluginID =key.split('|')
	sheet.cell(row=(current_ip),column=1).value = int(pluginID)
	sheet.cell(row=(current_ip),column=2).value = name
	sheet.cell(row=(current_ip),column=3).value = int(port)
	sheet.cell(row=(current_ip),column=4).value = protocol
	sheet.cell(row=(current_ip),column=5).value = name+" "+"("+protocol+"/"+port+")"
	sheet.cell(row=(current_ip),column=6).value = str(vulnerabilities[key]["svc_name"])
	if "plugin_output" in vulnerabilities[key]:
		#sheet.cell(row=(current_ip),column=8).value = vulnerabilities[key]["plugin_output"][0]
		if pluginID in ssl_cipher_plugin_ids:
			#print("\n\nIP : "+name)
			ciphers = ssl_cipher_finder.findall(vulnerabilities[key]["plugin_output"][0])
			all_cipher = ""
			for cipher_arr in ciphers:
				cipher = cipher_arr[0]
				if cipher is not None:
					all_cipher += cipher+"\n"
			sheet.cell(row=(current_ip),column=7).value = all_cipher
	current_ip += 1;
wb.save("ips.xlsx")	
